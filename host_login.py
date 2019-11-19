import codecs
import openpyxl as xl
import paramiko
username=""
mypass=""
f = xl.load_workbook("DECOM LIST lae.xlsx")
g = f["Sheet1"]
Hostname = [None]*3
j=0
for i in range(1,4):
    Hostname[j]=g.cell(i,1).value
    j+=1

i=1
for server in Hostname:
    try:
        #print(server)
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

        ssh.connect(server, username=username, password=mypass)
        ssh_stdout,ssh_stdin,ssh_stderr = ssh.exec_command("df -hP|wc -l")
        #ssh_stdout1,ssh_stdin1,ssh_stderr1 = ssh.exec_command("56611210")
        #ssh_stdout2,ssh_stdin2,ssh_stderr2 = ssh.exec_command("ls")
        about1 = (ssh_stdin.readlines())

        #about = (ssh_stdin1.readlines())
        #about2 = (ssh_stdin2.readlines())
        #print(f'### {server} \t {about1} \n')
        #print(f'### {server} \t {about} \n')

        print(f'### {server} \t {about1} ')
        g.cell(i, 3).value = (f'{server} \n {about1}')
        i+=1

    except:
        print(f'{server} \t "permission denied" ' )

        g.cell(i, 3).value = (f'{server} \t "permission denied" ')
        i+=1
        #print(f'{server} ---network issue')

        continue

f.save("DECOM LIST lae.xlsx")
